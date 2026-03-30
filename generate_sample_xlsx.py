from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.worksheet.datavalidation import DataValidation

def generate():
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Questions"

    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF")
    label_font = Font(bold=True)
    header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
    label_fill = PatternFill(start_color="F0F2F5", end_color="F0F2F5", fill_type="solid")
    
    # Word wrap applied everywhere
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap_align = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

    # --- Sheet Protection ---
    ws.protection.sheet = True

    # --- Section: Metadata (Rows 1-3) ---
    ws.column_dimensions['A'].width = 45 
    
    meta_labels = [
        ("Test Title *", "Biology Midterm - 2026"),
        ("Test Description (optional)", "Assessment focusing on Genetics."),
        ("Test Duration (minutes, optional)", "60")
    ]

    for idx, (label, val) in enumerate(meta_labels, 1):
        # Label
        cell_l = ws.cell(row=idx, column=1, value=label)
        cell_l.font = label_font
        cell_l.fill = label_fill
        cell_l.alignment = left_wrap_align
        cell_l.protection = Protection(locked=True)

        # Value (Editable)
        cell_v = ws.cell(row=idx, column=2, value=val)
        cell_v.protection = Protection(locked=False)
        cell_v.alignment = left_wrap_align
        ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=16)
        ws.row_dimensions[idx].height = 40 # Increased height for wrap

    # --- Table Headers (Row 5) ---
    headers = [
        "Question Type *", "Question Text *", "Correct Answer *", "Points *", 
        "Neg. Marks", "Time Limit (sec)", "Option A *", "Option B *", 
        "Option C", "Option D", "Option E", "Option F", "Option G", 
        "Option H", "Option I", "Option J"
    ]
    
    header_row = 5
    ws.row_dimensions[header_row].height = 30
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.protection = Protection(locked=True)
        
        # Adjust Widths
        if "Text" in header:
            ws.column_dimensions[cell.column_letter].width = 50
        elif "Option" in header:
             ws.column_dimensions[cell.column_letter].width = 20
        else:
             ws.column_dimensions[cell.column_letter].width = 18

    # --- Data Validations ---
    dv_qtype = DataValidation(type="list", formula1='"MCQ,Single Line"', allow_blank=False)
    ws.add_data_validation(dv_qtype)
    dv_qtype.add(f"A{header_row+1}:A500")

    dv_ans = DataValidation(type="list", formula1='"A,B,C,D,E,F,G,H,I,J"', allow_blank=False)
    dv_ans.showErrorMessage = False
    ws.add_data_validation(dv_ans)
    dv_ans.add(f"C{header_row+1}:C500")

    # Legend
    legend_cell = ws.cell(row=4, column=1, value="Note: * indicates mandatory fields. Word wrap enabled for clarity.")
    legend_cell.font = Font(italic=True, size=10, color="666666")
    legend_cell.protection = Protection(locked=True)

    # Unlock area
    for r in range(6, 501):
        ws.row_dimensions[r].height = 60 # Default height for question rows
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.protection = Protection(locked=False)
            cell.alignment = left_wrap_align

    wb.save("Swaya_me_Test_Template.xlsx")
    print("Swaya_me_Test_Template.xlsx updated with word wrap.")

if __name__ == "__main__":
    generate()
