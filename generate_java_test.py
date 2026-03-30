import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Protection

def populate():
    # Load template
    wb = openpyxl.load_workbook("Swaya_me_Test_Template.xlsx")
    ws = wb.active

    # --- Styles ---
    # Word wrap applied everywhere
    left_wrap_align = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

    # --- Set Metadata ---
    ws.cell(row=1, column=2, value="ISC Class 11 - Java Computer Science (Full Syllabus Review)")
    ws.cell(row=2, column=2, value="Sections: Fundamentals, Control Flow, Arrays, Strings, and Classes. Time limit recommended for 45 minutes.")
    ws.cell(row=3, column=2, value="45")
    
    # Metadata rows height adjusted
    for r in range(1, 4):
        ws.row_dimensions[r].height = 50

    # Section Headers (conceptual, just in question rows)
    # Col A: Type, B: Text, C: Ans, D: Pts, E: Neg, F: Time, G+: A-J
    questions = [
        # --- Section 1: Fundamentals & Data Types ---
        ["MCQ", "What is the return type of the hashCode() method in Java?", "B", 1, 0, 30, "String", "int", "boolean", "Object", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "Which of these is a valid declaration of a float variable in Java?", "A", 1, 0, 30, "float f = 1.2f;", "float f = 1.2;", "float f = 1.2D;", "float f = 1.2L;", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "What will be the output of: System.out.println(10 % 3); ?", "B", 1, 0, 20, "3", "1", "0.33", "Error", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "Which character is used to escape other characters in a Java String?", "C", 1, 0, 20, "Forward Slash (/)","Backtick (`)","Backslash (\\)","Plus (+)", "-", "-", "-", "-", "-", "-"],
        ["Single Line", "How many bits are used to represent a 'char' data type in Java?", "16", 1, 0, 30, "-", "-", "-", "-", "-", "-", "-", "-", "-", "-"],

        # --- Section 2: Flow Control & Code Snippets ---
        ["MCQ", "What will be the output of: for(int i=0; i<3; i++) { System.out.print(i); } ?", "A", 2, 0, 45, "012", "0123", "123", "Infinite Loop", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "Which loop is guaranteed to execute at least once?", "D", 1, 0, 20, "for", "while", "nested-for", "do-while", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "What will be the value of 'a' if: int a = 5; a = a++ + ++a;?", "C", 2, 1, 60, "10", "11", "12", "13", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "Which statement is used to skip the current iteration of a loop and move to the next one?", "B", 1, 0, 30, "break", "continue", "return", "exit", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "In a switch statement, which block is executed if no case value matches the expression?", "D", 1, 0, 20, "else", "otherwise", "catch", "default", "-", "-", "-", "-", "-", "-"],

        # --- Section 3: Strings & Math ---
        ["MCQ", "What is the value of: Math.pow(2, 3)?", "B", 1, 0, 20, "6.0", "8.0", "9.0", "1.0", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "What will: 'Hello World'.substring(6) return?", "C", 2, 0, 45, "Hello", "World ", "World", "Error", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "What is the output of: 'Java'.indexOf('a');?", "A", 1, 0, 30, "1", "2", "3", "0", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "Which method returns a String in lowercase?", "D", 1, 0, 30, "lowerCase()", "toLower()", "convertLower()", "toLowerCase()", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "What is the result of Math.ceil(4.2)?", "A", 1, 0, 20, "5.0", "4.0", "4.5", "6.0", "-", "-", "-", "-", "-", "-"],
        ["Single Line", "Name the built-in Math method used to find the square root of a number.", "Math.sqrt", 1, 0, 30, "-", "-", "-", "-", "-", "-", "-", "-", "-", "-"],

        # --- Section 4: Arrays ---
        ["MCQ", "How do you access the first element in an array named 'arr'?", "B", 1, 0, 20, "arr[1]", "arr[0]", "arr(0)", "arr.first()", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "What is the default value of an uninitialized int array element?", "A", 1, 0, 20, "0", "null", "false", "garbage", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "What will happen if you access an index beyond the array length?", "D", 1, 0, 30, "Prints null", "Returns 0", "Compile error", "Runtime exception", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "Which property is used to get the number of elements in an array?", "C", 1, 0, 20, "size", "count", "length", "getSize()", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "How many elements are in: int[][] table = new int[3][4]; ?", "B", 2, 0, 45, "7", "12", "3", "4", "-", "-", "-", "-", "-", "-"],

        # --- Section 5: OOP & Wrapper Classes ---
        ["MCQ", "An instance of a class is also known as a/an:", "C", 1, 0, 20, "Method", "Variable", "Object", "Package", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "Which keyword is used to refer to the current object?", "A", 1, 0, 20, "this", "super", "self", "instanceof", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "Which constructor has no parameters?", "B", 1, 0, 20, "Empty constructor", "Default constructor", "Static constructor", "Super constructor", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "Which wrapper class is used for the primitive type 'int'?", "D", 1, 0, 20, "Int", "IntWrapper", "IntClass", "Integer", "-", "-", "-", "-", "-", "-"],
        ["Single Line", "Keyword used to prevent a variable from being changed after initialization.", "final", 1, 0, 30, "-", "-", "-", "-", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "Which access specifier allows access only within the same package and its subclasses in other packages?", "B", 2, 0, 45, "private", "protected", "public", "default", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "What will 'Character.isDigit('7')' return?", "B", 1, 0, 30, "7", "true", "false", "Error", "-", "-", "-", "-", "-", "-"],
        ["MCQ", "Which method of Object class is called just before the garbage collector destroys the object?", "A", 2, 0, 45, "finalize()", "destroy()", "stop()", "delete()", "-", "-", "-", "-", "-", "-"],
        ["Single Line", "Name the operator used to create a new object in memory.", "new", 1, 0, 20, "-", "-", "-", "-", "-", "-", "-", "-", "-", "-"],
    ]

    # Write questions (A6 onwards)
    ws.delete_rows(6, ws.max_row)
    for row_idx, q_data in enumerate(questions, 6):
        ws.row_dimensions[row_idx].height = 80 # Multi-line wrap space
        for col_idx, val in enumerate(q_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.protection = Protection(locked=False)
            cell.alignment = left_wrap_align

    wb.save("ISC_Class_11_Java_Comprehensive_45m.xlsx")
    print("ISC_Class_11_Java_Comprehensive_45m.xlsx generated with 30 questions + word wrap.")

if __name__ == "__main__":
    populate()
