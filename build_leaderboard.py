#!/usr/bin/env python3
"""
Build folder-9 exam leaderboard Excel for meetnishant@gmail.com.
5 tabs: B7 (quiz185/sess442), B8 (quiz186/sess438),
        C9 (quiz200/sess444), C10 (quiz201/sess448+436), C11 (quiz202/sess435)
"""
import sys
sys.path.insert(0, '/home/vinay/Swaya.me/backend')

import pymysql
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime

# ── DB connection ────────────────────────────────────────────────────────────
DB = dict(host='127.0.0.1', user='swayame_user',
          password='Sw4y4m3_S3cur3_P4ssw0rd!2026', db='swayame',
          charset='utf8mb4')

def conn():
    return pymysql.connect(**DB, cursorclass=pymysql.cursors.DictCursor)

# ── Quiz → (tab_name, session_ids, quiz_id) ──────────────────────────────────
EXAMS = [
    ('B7',  185, [442],      2700),
    ('B8',  186, [438],      3000),
    ('C9',  200, [444],      2700),
    ('C10', 201, [448, 436], 3600),
    ('C11', 202, [435],      3600),
]

def fetch_questions(cur, quiz_id):
    cur.execute(
        "SELECT id, `order`, points, negative_points "
        "FROM questions WHERE quiz_id=%s AND question_type='mcq' ORDER BY `order`",
        (quiz_id,)
    )
    return cur.fetchall()

def fetch_participants(cur, session_ids):
    ph = ','.join(['%s'] * len(session_ids))
    cur.execute(
        f"SELECT id, session_id, display_name, email, started_at, completed_at, is_abandoned "
        f"FROM participants WHERE session_id IN ({ph})",
        session_ids
    )
    return cur.fetchall()

def fetch_answers(cur, session_ids):
    ph = ','.join(['%s'] * len(session_ids))
    cur.execute(
        f"SELECT participant_id, question_id, is_correct "
        f"FROM answers WHERE session_id IN ({ph})",
        session_ids
    )
    return cur.fetchall()

def fetch_integrity(cur, quiz_id, participant_ids):
    if not participant_ids:
        return {}
    ph = ','.join(['%s'] * len(participant_ids))
    cur.execute(
        f"SELECT participant_id, integrity_score, violation_count, is_locked "
        f"FROM proctoring_sessions WHERE quiz_id=%s AND participant_id IN ({ph})",
        [quiz_id] + list(participant_ids)
    )
    rows = cur.fetchall()
    # pick best (highest integrity) proctoring row per participant
    result = {}
    for r in rows:
        pid = r['participant_id']
        if pid not in result or r['integrity_score'] > result[pid]['integrity_score']:
            result[pid] = r
    return result

def compute_leaderboard(questions, participants, answers):
    ans_map = {}  # {pid: {qid: is_correct}}
    for a in answers:
        ans_map.setdefault(a['participant_id'], {})[a['question_id']] = a['is_correct']

    max_score = sum(q['points'] for q in questions)
    entries = []
    for p in participants:
        if not p['completed_at'] or p['is_abandoned']:
            continue
        p_ans = ans_map.get(p['id'], {})
        score = 0
        correct = 0
        for q in questions:
            ic = p_ans.get(q['id'])
            if ic:          # is_correct == 1 (pymysql returns int, not bool)
                score += q['points']
                correct += 1
            elif ic == 0 and ic is not None:
                score -= q['negative_points']
        score = max(0, score)
        time_taken = None
        if p['started_at'] and p['completed_at']:
            time_taken = (p['completed_at'] - p['started_at']).total_seconds()
        entries.append({
            'pid': p['id'],
            'name': p['display_name'] or '—',
            'email': p['email'] or '',
            'score': score,
            'max_score': max_score,
            'correct': correct,
            'total_q': len(questions),
            'time_taken': time_taken,
            'completed_at': p['completed_at'],
        })

    # Sort: score desc, time asc
    entries.sort(key=lambda x: (-x['score'], x['time_taken'] or 9_999_999))
    for i, e in enumerate(entries, 1):
        e['rank'] = i
    return entries, max_score

def fmt_time(seconds):
    if seconds is None:
        return '—'
    m, s = divmod(int(seconds), 60)
    h, m = divmod(m, 60)
    if h:
        return f'{h}h {m}m {s}s'
    return f'{m}m {s}s'

# ── Style helpers ─────────────────────────────────────────────────────────────
HEADER_BG   = '1F3864'  # dark navy
SUBHDR_BG   = '2E75B6'  # mid blue
GOLD        = 'FFD700'
SILVER      = 'C0C0C0'
BRONZE      = 'CD7F32'
RANK_COLORS = {1: GOLD, 2: SILVER, 3: BRONZE}

GREEN_BG    = 'E2EFDA'
RED_BG      = 'FFDCE1'
STRIPE_ODD  = 'DDEEFF'
STRIPE_EVEN = 'FFFFFF'
CROSS_BG    = 'FFF2CC'  # yellow tint for cross-exam note

thin = Side(style='thin', color='BFBFBF')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font(bold=True, size=11, color='FFFFFF'):
    return Font(name='Calibri', bold=bold, size=size, color=color)

def cell_font(bold=False, size=10, color='000000', italic=False):
    return Font(name='Calibri', bold=bold, size=size, color=color, italic=italic)

def apply_fill(cell, hex_color):
    cell.fill = PatternFill('solid', fgColor=hex_color)

def style_header_row(ws, row, ncols, bg=HEADER_BG):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill('solid', fgColor=bg)
        c.font = hdr_font()
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER

def write_exam_sheet(ws, tab_name, entries, max_score, integrity_map, cross_map):
    """Write one leaderboard sheet."""
    # ── Title block ──────────────────────────────────────────────────────────
    COLS = 8  # Rank | Name | Email | Score | % | Correct | Time | Integrity
    ws.merge_cells(f'A1:{get_column_letter(COLS)}1')
    title = ws['A1']
    title.value = f'Exam {tab_name} — Leaderboard'
    title.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    title.alignment = Alignment(horizontal='center', vertical='center')
    title.fill = PatternFill('solid', fgColor=HEADER_BG)
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f'A2:{get_column_letter(COLS)}2')
    sub = ws['A2']
    sub.value = (f'Max Score: {max_score}  |  '
                 f'Completed: {len(entries)}  |  '
                 f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")}')
    sub.font = Font(name='Calibri', size=10, color='FFFFFF', italic=True)
    sub.alignment = Alignment(horizontal='center', vertical='center')
    sub.fill = PatternFill('solid', fgColor=SUBHDR_BG)
    ws.row_dimensions[2].height = 18

    # ── Column headers ────────────────────────────────────────────────────────
    headers = ['Rank', 'Name', 'Email', f'Score\n(/{max_score})',
               'Score %', f'Correct\n(/{entries[0]["total_q"] if entries else "?"})',
               'Time Taken', 'Integrity\nScore']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, COLS)
    ws.row_dimensions[3].height = 32

    # ── Column widths ────────────────────────────────────────────────────────
    widths = [7, 28, 36, 10, 10, 10, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Data rows ────────────────────────────────────────────────────────────
    for idx, e in enumerate(entries):
        row = idx + 4
        pid = e['pid']
        integ = integrity_map.get(pid, {})
        integ_score = integ.get('integrity_score', '—')
        is_locked   = integ.get('is_locked', 0)
        violations  = integ.get('violation_count', 0)

        cross_note = cross_map.get(e['email'], '')

        display_name = e['name']
        if cross_note:
            display_name = f"{e['name']}  ⟨{cross_note}⟩"

        rank = e['rank']
        stripe = STRIPE_ODD if idx % 2 == 0 else STRIPE_EVEN

        row_data = [
            rank,
            display_name,
            e['email'],
            e['score'],
            f"{round(e['score']/max_score*100 if max_score else 0, 1)}%",
            e['correct'],
            fmt_time(e['time_taken']),
            integ_score if isinstance(integ_score, int) else '—',
        ]

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = BORDER
            c.alignment = Alignment(horizontal='center' if col not in (2, 3) else 'left',
                                    vertical='center', wrap_text=(col in (2,)))

            # Base stripe
            bg = stripe
            if rank in RANK_COLORS and col == 1:
                bg = RANK_COLORS[rank]
            elif cross_note:
                bg = CROSS_BG
            elif is_locked:
                bg = 'FFE0E0'

            apply_fill(c, bg)
            bold = rank <= 3
            c.font = cell_font(bold=bold, size=10,
                               italic=(col == 2 and bool(cross_note)))

        # Rank medal styling
        rank_cell = ws.cell(row=row, column=1)
        if rank in RANK_COLORS:
            rank_cell.font = Font(name='Calibri', bold=True, size=11, color='000000')

        # Integrity color coding
        integ_cell = ws.cell(row=row, column=8)
        if isinstance(integ_score, int):
            if integ_score >= 80:
                apply_fill(integ_cell, 'C6EFCE')
                integ_cell.font = cell_font(color='276221', bold=(rank <= 3))
            elif integ_score >= 50:
                apply_fill(integ_cell, 'FFEB9C')
                integ_cell.font = cell_font(color='9C6500', bold=(rank <= 3))
            else:
                apply_fill(integ_cell, 'FFC7CE')
                integ_cell.font = cell_font(color='9C0006', bold=(rank <= 3))
        if is_locked:
            integ_cell.value = f"{integ_score} 🔒" if isinstance(integ_score, int) else '🔒'

        ws.row_dimensions[row].height = 18

    # ── Freeze top rows ───────────────────────────────────────────────────────
    ws.freeze_panes = 'A4'

    # ── Legend row ────────────────────────────────────────────────────────────
    legend_row = len(entries) + 5
    ws.merge_cells(f'A{legend_row}:{get_column_letter(COLS)}{legend_row}')
    leg = ws[f'A{legend_row}']
    leg.value = ('Legend:  ⟨Also appeared in X⟩ = candidate sat multiple tests  |  '
                 '🔒 = exam locked by proctoring  |  '
                 'Integrity: Green ≥80, Yellow 50–79, Red <50')
    leg.font = Font(name='Calibri', size=9, italic=True, color='595959')
    leg.alignment = Alignment(horizontal='left', vertical='center')
    apply_fill(leg, 'F2F2F2')
    ws.row_dimensions[legend_row].height = 16

# ── Summary sheet ─────────────────────────────────────────────────────────────
def write_summary_sheet(ws, all_email_tabs):
    ws.title = 'Summary'
    ws.sheet_view.showGridLines = False

    ws.merge_cells('A1:E1')
    t = ws['A1']
    t.value = 'Folder-9 Exam — Candidate Summary'
    t.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    t.alignment = Alignment(horizontal='center', vertical='center')
    t.fill = PatternFill('solid', fgColor=HEADER_BG)
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:E2')
    s = ws['A2']
    s.value = 'Candidates who appeared in more than one test are highlighted.'
    s.font = Font(name='Calibri', size=10, italic=True, color='FFFFFF')
    s.alignment = Alignment(horizontal='center', vertical='center')
    s.fill = PatternFill('solid', fgColor=SUBHDR_BG)
    ws.row_dimensions[2].height = 18

    hdrs = ['Email', 'Name', 'Tests Appeared', 'Test Tabs', 'Cross-Exam Note']
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill = PatternFill('solid', fgColor=SUBHDR_BG)
        c.font = hdr_font()
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = BORDER
    ws.row_dimensions[3].height = 22

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 36

    multi = {e: tabs for e, tabs in all_email_tabs.items() if len(tabs) > 1}
    single = {e: tabs for e, tabs in all_email_tabs.items() if len(tabs) == 1}

    row = 4
    for email, tabs in sorted(multi.items()):
        name = tabs[0][1]
        tab_names = [t[0] for t in tabs]
        cross = 'Also appeared in ' + ', '.join(t for t in tab_names[1:])
        for col, val in enumerate([email, name, len(tabs),
                                    ', '.join(tab_names), cross], 1):
            c = ws.cell(row=row, column=col, value=val)
            apply_fill(c, CROSS_BG)
            c.font = cell_font(bold=True, size=10)
            c.alignment = Alignment(horizontal='left' if col in (1,2,4,5) else 'center',
                                    vertical='center')
            c.border = BORDER
        ws.row_dimensions[row].height = 16
        row += 1

    if multi:
        row += 1  # gap

    for email, tabs in sorted(single.items()):
        name = tabs[0][1]
        tab_names = [t[0] for t in tabs]
        for col, val in enumerate([email, name, 1,
                                    tab_names[0], '—'], 1):
            c = ws.cell(row=row, column=col, value=val)
            stripe = STRIPE_ODD if (row % 2 == 0) else STRIPE_EVEN
            apply_fill(c, stripe)
            c.font = cell_font(size=10)
            c.alignment = Alignment(horizontal='left' if col in (1,2,4,5) else 'center',
                                    vertical='center')
            c.border = BORDER
        ws.row_dimensions[row].height = 16
        row += 1

    ws.freeze_panes = 'A4'

# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    all_email_tabs = {}  # {email: [(tab_name, display_name), ...]}

    # first pass: collect all data
    exam_data = []
    with conn() as db:
        cur = db.cursor()
        for tab_name, quiz_id, session_ids, duration in EXAMS:
            questions   = fetch_questions(cur, quiz_id)
            participants= fetch_participants(cur, session_ids)
            answers     = fetch_answers(cur, session_ids)
            entries, max_score = compute_leaderboard(questions, participants, answers)
            pids = [e['pid'] for e in entries]
            integrity_map = fetch_integrity(cur, quiz_id, pids)
            exam_data.append((tab_name, entries, max_score, integrity_map))

            for e in entries:
                em = (e['email'] or '').lower().strip()
                if em:
                    tabs_list = all_email_tabs.setdefault(em, [])
                    # Record each tab only once — multiple attempts in the same
                    # test don't count as "also appeared in" a different test
                    if not any(t[0] == tab_name for t in tabs_list):
                        tabs_list.append((tab_name, e['name']))

    # second pass: build cross-exam notes then write sheets
    for tab_name, entries, max_score, integrity_map in exam_data:
        ws = wb.create_sheet(title=tab_name)
        ws.sheet_view.showGridLines = False

        cross_map = {}
        for e in entries:
            em = (e['email'] or '').lower().strip()
            tabs = all_email_tabs.get(em, [])
            if len(tabs) > 1:
                other_tabs = [t[0] for t in tabs if t[0] != tab_name]
                if other_tabs:
                    cross_map[e['email']] = 'Also in ' + ', '.join(other_tabs)

        write_exam_sheet(ws, tab_name, entries, max_score, integrity_map, cross_map)
        print(f'  {tab_name}: {len(entries)} candidates')

    # summary sheet
    ws_sum = wb.create_sheet(title='Summary', index=0)
    write_summary_sheet(ws_sum, all_email_tabs)

    out = '/home/vinay/Swaya.me/folder9_leaderboard.xlsx'
    wb.save(out)
    print(f'\nSaved → {out}')

if __name__ == '__main__':
    main()
