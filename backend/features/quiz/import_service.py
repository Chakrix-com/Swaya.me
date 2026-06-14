"""
Excel Import/Export Service - Logic for handling bulk quiz data via XLSX
"""
import io
from shared.utils.html_sanitizer import sanitize_html, sanitize_plain
from pathlib import Path
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta

from persistence.models.quiz import (
    Quiz, Question, QuizStatus, QuizType, QuestionType
)
from persistence.models.core import Event
from features.quiz.schemas import QuizCreate, QuestionCreate
from shared.exceptions.quiz import QuizValidationError
from fastapi import HTTPException

class ExcelImportService:
    """Service to handle Excel import/export for quizzes"""

    # --- Localized Labels & Synonyms ---
    EXCEL_TRANSLATIONS = {
        "en": {
            "meta_title": "Title *", "meta_desc": "Description", "meta_dur": "Duration (m)", "meta_type": "Quiz Type *",
            "quiz": "Quiz", "test": "Test",
            "col_type": "Type *", "col_text": "Question Text *", "col_ans": "Correct Answer *", 
            "col_points": "Points", "col_neg": "Neg. Marks", "col_time": "Time (sec)",
            "mcq": "MCQ", "single_line": "Single Line"
        },
        "de": {
            "meta_title": "Titel *", "meta_desc": "Beschreibung", "meta_dur": "Dauer (m)", "meta_type": "Quiz-Typ *",
            "quiz": "Quiz", "test": "Test",
            "col_type": "Typ *", "col_text": "Fragetext *", "col_ans": "Korrekte Antwort *", 
            "col_points": "Punkte", "col_neg": "Negativpunkte", "col_time": "Zeit (Sek.)",
            "mcq": "MCQ", "single_line": "Einzeiler"
        },
        "fr": {
            "meta_title": "Titre *", "meta_desc": "Description", "meta_dur": "Durée (m)", "meta_type": "Type de quiz *",
            "quiz": "Quiz", "test": "Test",
            "col_type": "Type *", "col_text": "Texte de la question *", "col_ans": "Réponse correcte *", 
            "col_points": "Points", "col_neg": "Points négatifs", "col_time": "Temps (sec)",
            "mcq": "QCM", "single_line": "Ligne unique"
        },
        "hi": {
            "meta_title": "शीर्षक *", "meta_desc": "विवरण", "meta_dur": "अवधि (m)", "meta_type": "क्विज़ का प्रकार *",
            "quiz": "क्विज़", "test": "परीक्षा",
            "col_type": "प्रकार *", "col_text": "प्रश्न पाठ *", "col_ans": "सही उत्तर *", 
            "col_points": "अंक", "col_neg": "नकारात्मक अंक", "col_time": "समय (sec)",
            "mcq": "MCQ", "single_line": "एक पंक्ति"
        }
    }

    def _get_localized(self, lang: str = "en") -> Dict[str, str]:
        # Handle variants like 'en-US'
        short_lang = (lang or "en").split("-")[0].lower()
        return self.EXCEL_TRANSLATIONS.get(short_lang, self.EXCEL_TRANSLATIONS["en"])

    @staticmethod
    def get_template_path() -> str:
        return str(Path(__file__).parents[3] / "Swaya_me_Test_Template.xlsx")

    @staticmethod
    def _char_to_index(char: str) -> Optional[int]:
        if not char or len(char) != 1:
            return None
        char = char.upper()
        if 'A' <= char <= 'J':
            return ord(char) - ord('A')
        return None

    @staticmethod
    def _index_to_char(index: int) -> str:
        if 0 <= index <= 9:
            return chr(ord('A') + index)
        return ""

    async def parse_excel(self, file_content: bytes) -> Dict[str, Any]:
        """Parse the XLSX file and return raw data for validation"""
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=501, detail="Excel processing library not available")

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            ws = wb.active

            # --- Extract Metadata (Rows 1-4) ---
            title = ws["B1"].value
            description = ws["B2"].value
            duration = ws["B3"].value
            quiz_type_label = str(ws["B4"].value or "").strip()
            
            # Map labels back to internal values
            quiz_type = "exam" if quiz_type_label in ["Test", "परीक्षा", "Exam"] else "quiz"

            questions = []
            # Start from row 6
            for row in ws.iter_rows(min_row=6, values_only=True):
                if not any(row): continue
                
                # Broaden type recognition for localization
                raw_type = str(row[0] or "").strip()
                q_type = "mcq"
                if any(variant in raw_type for variant in ["Single", "Ligne", "Einzeiler", "एक पंक्ति"]):
                    q_type = "single_line"
                
                questions.append({
                    "type": q_type,
                    "text": row[1],
                    "answer": str(row[2]) if row[2] is not None else "",
                    "points": row[3] if row[3] is not None else 1,
                    "neg_marks": row[4] if row[4] is not None else 0,
                    "time_limit": row[5] if row[5] is not None else 30,
                    "options": [str(x) for x in row[6:16] if x is not None and x != "-"]
                })

            return {
                "title": title,
                "description": description,
                "duration_minutes": duration,
                "quiz_type": quiz_type,
                "questions": questions
            }
        except Exception as e:
            raise QuizValidationError(f"Failed to parse Excel file: {str(e)}")

    def validate_import(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Validate parsed data and return structured preview with errors"""
        errors = []
        validated_questions = []
        
        if not data.get("title"):
            errors.append("Quiz Title is mandatory (Cell B1)")

        for idx, q in enumerate(data.get("questions", []), 1):
            row_errors = []
            
            # Text Validation
            if not q["text"]:
                row_errors.append("Question Text is mandatory")

            # Answer Validation
            correct_index = None
            options = q["options"]
            
            if q["type"] == "mcq":
                if len(options) < 2:
                    row_errors.append("MCQ requires at least 2 options")
                
                correct_index = self._char_to_index(q["answer"])
                if correct_index is None:
                    row_errors.append(f"Invalid correct option: {q['answer']}. Use A-J.")
                elif correct_index >= len(options):
                    row_errors.append(f"Option {q['answer']} is selected, but only {len(options)} options provided.")
            
            elif q["type"] == "single_line":
                if not q["answer"] or q["answer"] == "-":
                    row_errors.append("Expected Answer is mandatory for Single Line questions")
            
            validated_questions.append({
                "index": idx,
                "type": "MCQ" if q["type"] == "mcq" else "Single Line",
                "text": q["text"],
                "answer": q["answer"],
                "points": q["points"],
                "neg_marks": q["neg_marks"],
                "time_limit": q["time_limit"],
                "options": q["options"],
                "errors": row_errors
            })
            if row_errors:
                errors.append(f"Row {idx+5}: " + ", ".join(row_errors))

        return {
            "title": data.get("title"),
            "description": data.get("description"),
            "quiz_type": data.get("quiz_type"),
            "duration_minutes": data.get("duration_minutes"),
            "questions": validated_questions,
            "errors": errors,
            "canImport": len(errors) == 0
        }

    async def create_from_import(self, db: AsyncSession, data: Dict[str, Any], current_user: Any) -> Quiz:
        """Persist the imported quiz and questions to database"""
        # Prefix title with type to help host distinguish
        quiz_type_val = data.get("quiz_type", "exam")
        type_prefix = "Exam" if quiz_type_val == "exam" else "Quiz"
        
        # Create Event
        event = Event(
            tenant_id=current_user.tenant_id,
            creator_id=current_user.user_id,
            title=f"{type_prefix} - {data['title']}",
            description=data.get("description"),
            join_code=None
        )
        db.add(event)
        await db.flush()

        # Create Quiz
        try:
            quiz_type = QuizType(quiz_type_val)
        except ValueError:
            quiz_type = QuizType.EXAM

        quiz = Quiz(
            tenant_id=current_user.tenant_id,
            event_id=event.id,
            title=data["title"],
            description=data["description"],
            quiz_type=quiz_type,
            status=QuizStatus.DRAFT,
            # For exams, set windows and limits
            exam_start_at=datetime.utcnow() if quiz_type == QuizType.EXAM else None,
            exam_end_at=datetime.utcnow() + timedelta(days=1) if quiz_type == QuizType.EXAM else None,
            exam_time_limit_seconds=int(data["duration_minutes"] or 0) * 60 if data.get("duration_minutes") and quiz_type == QuizType.EXAM else None
        )
        db.add(quiz)
        await db.flush()
        await db.refresh(quiz)

        # Create Questions
        for idx, q in enumerate(data["questions"]):
            # validated questions use "MCQ"/"Single Line"; raw use "mcq"/"single_line"
            raw_type = str(q.get("type", "MCQ")).strip().upper()
            q_type = QuestionType.MCQ if raw_type in ("MCQ", "MCQ") else QuestionType.SINGLE_LINE
            if "SINGLE" in raw_type or "LINE" in raw_type:
                q_type = QuestionType.SINGLE_LINE

            db_options = []
            correct_index = None

            if q_type == QuestionType.MCQ:
                db_options = q["options"]
                correct_index = self._char_to_index(q["answer"])
            else:
                db_options = [q["answer"]]
                correct_index = None

            question = Question(
                quiz_id=quiz.id,
                question_type=q_type,
                text=sanitize_html(q["text"]),
                order=idx,
                options=[str(o).strip() for o in db_options],
                correct_answer_index=correct_index,
                points=float(q["points"] or 1),
                negative_points=float(q["neg_marks"] if q["neg_marks"] is not None else 0),
                max_time_seconds=int(q["time_limit"] or 30)
            )
            db.add(question)

        await db.commit()
        await db.refresh(quiz)
        return quiz

    def generate_excel_from_draft(self, draft_data: Dict[str, Any]) -> bytes:
        """Generate a populated XLSX from frontend draft JSON"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Protection
            from openpyxl.worksheet.datavalidation import DataValidation
        except ImportError:
            raise HTTPException(status_code=501, detail="Excel processing library not available")

        wb = Workbook()
        ws = wb.active
        ws.title = "Swaya Quiz Draft"

        # --- Localized Labels ---
        lang = draft_data.get("lang", "en")
        L = self._get_localized(lang)

        # --- Styles ---
        header_font = Font(bold=True, color="FFFFFF")
        label_font = Font(bold=True)
        header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
        label_fill = PatternFill(start_color="F0F2F5", end_color="F0F2F5", fill_type="solid")
        left_wrap_align = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.protection.sheet = True
        ws.column_dimensions['A'].width = 15

        # Metadata
        current_type = draft_data.get("quiz_type", "exam")
        meta_labels = [
            (L["meta_title"], draft_data.get("title", "")),
            (L["meta_desc"], draft_data.get("description", "")),
            (L["meta_dur"], draft_data.get("duration_minutes", "")),
            (L["meta_type"], L["test"] if current_type == "exam" else L["quiz"])
        ]

        for idx, (label, val) in enumerate(meta_labels, 1):
            ws.cell(row=idx, column=1, value=label).font = label_font
            ws.cell(row=idx, column=1).fill = label_fill
            ws.cell(row=idx, column=1).alignment = left_wrap_align
            
            cell_b = ws.cell(row=idx, column=2, value=val)
            cell_b.protection = Protection(locked=False)
            cell_b.alignment = left_wrap_align
            
            if idx <= 2:
                ws.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=16)
            
            ws.row_dimensions[idx].height = 25

        # Headers
        headers = [
            L["col_type"], L["col_text"], L["col_ans"], L["col_points"], 
            L["col_neg"], L["col_time"], "Option A", "Option B", 
            "Option C", "Option D", "Option E", "Option F", "Option G", 
            "Option H", "Option I", "Option J"
        ]
        header_row = 5
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            ws.column_dimensions[cell.column_letter].width = 40 if "Text" in header else 15

        # Data rows
        questions = draft_data.get("questions", [])
        for r_idx, q in enumerate(questions, 6):
            q_type = L["mcq"] if q.get("type", "mcq") == "mcq" else L["single_line"]
            text = q.get("text", "")
            
            ans = ""
            if (q.get("type", "mcq") == "mcq"):
                ans = self._index_to_char(q.get("correct_answer_index", 0))
            else:
                ans = q.get("expected_answer", "") or q.get("answer", "")

            row_data = [
                q_type, text, ans, q.get("points", 1), q.get("negative_points", 0), q.get("max_time_seconds", 30)
            ]
            
            # Options
            q_options = q.get("options", [])
            for i in range(10):
                row_data.append(q_options[i] if i < len(q_options) else "-")

            ws.row_dimensions[r_idx].height = 40
            for c_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.protection = Protection(locked=False)
                cell.alignment = left_wrap_align

        # Validations
        formula_qtype = f'"{L["mcq"]},{L["single_line"]}"'
        dv_qtype = DataValidation(type="list", formula1=formula_qtype, allow_blank=False)
        ws.add_data_validation(dv_qtype)
        dv_qtype.add(f"A6:A500")

        # Dropdown for correct answers A-J
        dv_ans = DataValidation(type="list", formula1='"A,B,C,D,E,F,G,H,I,J"', allow_blank=False)
        ws.add_data_validation(dv_ans)
        dv_ans.add(f"C6:C500")

        formula_type = f'"{L["quiz"]},{L["test"]}"'
        dv_type = DataValidation(type="list", formula1=formula_type, allow_blank=False)
        ws.add_data_validation(dv_type)
        dv_type.add(f"B4")

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
